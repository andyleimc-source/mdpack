"""XLSX → Markdown (one section per sheet)."""

from __future__ import annotations

from pathlib import Path

from ..utils import md_escape_cell
from .base import ConversionError, ConvertResult


def _format_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


class XlsxConverter:
    name = "xlsx"
    extensions = (".xlsx",)

    def convert(self, src: Path) -> ConvertResult:
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise ConversionError(
                src, "openpyxl is not installed — run: pip install 'mdpack[xlsx]'"
            ) from e

        try:
            wb = load_workbook(src, read_only=True, data_only=True)
        except Exception as e:
            raise ConversionError(src, f"failed to open xlsx: {e}") from e

        lines: list[str] = [f"# {src.stem}\n"]
        warnings: list[str] = []
        non_empty_sheets = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = [
                [_format_cell(c) for c in row]
                for row in ws.iter_rows(values_only=True)
            ]
            rows = [r for r in rows if any(cell.strip() for cell in r)]
            if not rows:
                warnings.append(f"sheet {sheet_name!r} is empty")
                continue

            width = max(len(r) for r in rows)
            rows = [r + [""] * (width - len(r)) for r in rows]
            header, data = rows[0], rows[1:]

            lines.append(f"\n## {sheet_name}\n")
            lines.append("| " + " | ".join(md_escape_cell(c) for c in header) + " |")
            lines.append("|" + "|".join(["---"] * width) + "|")
            for row in data:
                lines.append("| " + " | ".join(md_escape_cell(c) for c in row) + " |")
            non_empty_sheets += 1

        wb.close()

        if non_empty_sheets == 0:
            return ConvertResult(body="", title=src.stem, warnings=["all sheets empty"])

        return ConvertResult(body="\n".join(lines) + "\n", title=src.stem, warnings=warnings)
